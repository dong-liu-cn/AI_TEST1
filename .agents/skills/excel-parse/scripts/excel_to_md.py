#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
excel_to_md.py
Excel業務機能設計書を解析し、Markdown基本設計書に変換する共通スクリプト。

使用方法:
    py excel_to_md.py --input "path/to/file.xlsx" [--output "output.md"]

依存:
    pip install openpyxl lxml
"""

import argparse
import io
import os
import re
import sys
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# UTF-8出力修正
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
if sys.stderr.encoding and sys.stderr.encoding.lower() not in ("utf-8", "utf8"):
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

try:
    import openpyxl
except ImportError:
    print("openpyxlが必要です: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    from lxml import etree
except ImportError:
    etree = None  # 図形抽出時のみ必要


# =====================================================================
#  シート分類キーワード
# =====================================================================
SHEET_CATEGORIES = [
    ("change_history",      ["変更履歴"]),
    ("spec_overview",       ["仕様概要", "機能概要"]),
    ("design_overview",     ["設計概要"]),
    ("screen_items",        ["画面項目定義"]),
    ("screen_state",        ["画面状態定義", "画面状態"]),
    ("screen_check",        ["チェック定義", "チェックマトリクス", "チェックマトリックス"]),
    ("screen_source",       ["取得元定義", "画面取得元"]),
    ("screen_transition",   ["画面遷移図", "機能遷移図"]),
    ("excel_transfer",      ["移送表"]),
    ("excel_layout",        ["レイアウト"]),
    ("process_detail",      ["処理詳細", "処理設計"]),
    ("table_mapping",       ["テーブル項目", "マッピング"]),
]


def classify_sheets(sheet_names: List[str]) -> Dict[str, List[str]]:
    """シート名をカテゴリに分類する。"""
    classified = {}
    used = set()
    for cat, keywords in SHEET_CATEGORIES:
        matches = []
        for name in sheet_names:
            if name in used:
                continue
            for kw in keywords:
                if kw in name:
                    matches.append(name)
                    used.add(name)
                    break
        if matches:
            classified[cat] = matches
    # 未分類
    unclassified = [n for n in sheet_names if n not in used]
    if unclassified:
        classified["other"] = unclassified
    return classified


# =====================================================================
#  ユーティリティ
# =====================================================================
def cell_value(cell) -> str:
    """セル値を文字列に変換する。"""
    if cell is None or cell.value is None:
        return ""
    v = cell.value
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v == int(v):
            return str(int(v))
        return str(v)
    return str(v).strip()


def row_values(row) -> List[str]:
    """行のセル値リストを返す。"""
    return [cell_value(c) for c in row]


def is_empty_row(values: List[str]) -> bool:
    """全セルが空かどうか。"""
    return all(v == "" for v in values)


def trim_trailing_empty(values: List[str]) -> List[str]:
    """末尾の空セルを除去。"""
    while values and values[-1] == "":
        values.pop()
    return values


def escape_md(text: str) -> str:
    """Markdownテーブル内のパイプ文字・改行をエスケープする。"""
    text = text.replace("|", "\\|")
    text = text.replace("\n", "<br/>")
    text = text.replace("\r", "")
    return text


def make_md_table(headers: List[str], rows: List[List[str]]) -> str:
    """Markdownテーブルを生成する。"""
    lines = []
    h = [escape_md(h) for h in headers]
    lines.append("| " + " | ".join(h) + " |")
    lines.append("|" + "|".join(["---" for _ in headers]) + "|")
    for row in rows:
        # 列数を揃える
        padded = row + [""] * (len(headers) - len(row))
        r = [escape_md(v) for v in padded[:len(headers)]]
        lines.append("| " + " | ".join(r) + " |")
    return "\n".join(lines)


# =====================================================================
#  基本情報抽出
# =====================================================================
def extract_basic_info(wb: openpyxl.Workbook, sheet_names: List[str]) -> Dict[str, str]:
    """仕様概要シートから基本情報を抽出する。"""
    info = {}
    target = None
    for name in sheet_names:
        if "仕様概要" in name or "機能概要" in name:
            target = name
            break
    if not target:
        return info

    ws = wb[target]
    # 先頭20行をスキャンして基本情報を探す
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=False):
        vals = row_values(row)
        for i, v in enumerate(vals):
            if "管理No" in v or "管理ＮＯ" in v:
                # 次の非空セルが値
                for j in range(i + 1, len(vals)):
                    if vals[j]:
                        info["管理No"] = vals[j]
                        break
            elif "業務機能名" in v or "機能名" in v:
                for j in range(i + 1, len(vals)):
                    if vals[j]:
                        info["業務機能名"] = vals[j]
                        break
            elif "機能概要" in v and "機能概要" not in info:
                for j in range(i + 1, len(vals)):
                    if vals[j]:
                        info["機能概要"] = vals[j]
                        break
    return info


# =====================================================================
#  各シート種別の解析
# =====================================================================

def parse_change_history(wb: openpyxl.Workbook, sheet_name: str) -> str:
    """変更履歴シートをMarkdownに変換する。"""
    ws = wb[sheet_name]
    lines = [f"## 変更履歴\n"]
    
    headers = None
    data_rows = []
    for row in ws.iter_rows(min_row=1, values_only=False):
        vals = row_values(row)
        if is_empty_row(vals):
            continue
        vals = trim_trailing_empty(vals)
        if not vals:
            continue
        # ヘッダー行を検出（"No"を含む行）
        if headers is None and any("No" in v for v in vals):
            headers = vals
            continue
        if headers:
            data_rows.append(vals)
    
    if headers and data_rows:
        lines.append(make_md_table(headers, data_rows))
    else:
        lines.append("> ※ 変更履歴データが見つかりませんでした。")
    
    return "\n".join(lines)


def parse_spec_overview(wb: openpyxl.Workbook, sheet_name: str) -> str:
    """仕様概要シートをMarkdownに変換する。"""
    ws = wb[sheet_name]
    lines = [f"## 仕様概要\n"]
    
    # 基本情報
    basic_info = {}
    action_headers = None
    action_rows = []
    in_action_table = False
    
    for row in ws.iter_rows(min_row=1, values_only=False):
        vals = row_values(row)
        if is_empty_row(vals):
            continue
        
        # 基本情報キーの検出
        for i, v in enumerate(vals):
            if v in ("管理No", "管理ＮＯ"):
                for j in range(i + 1, len(vals)):
                    if vals[j]:
                        basic_info["管理No"] = vals[j]
                        break
            elif v in ("業務機能名",):
                for j in range(i + 1, len(vals)):
                    if vals[j]:
                        basic_info["業務機能名"] = vals[j]
                        break
            elif v == "機能概要":
                for j in range(i + 1, len(vals)):
                    if vals[j]:
                        basic_info["機能概要"] = vals[j]
                        break
        
        # 仕様概要テーブル（処理パターンテーブル）
        if any("処理パターン" in v for v in vals) or any("処理名" in v for v in vals):
            in_action_table = True
            # 次の行と結合してヘッダーを構築する可能性がある
            if action_headers is None:
                action_headers = vals
            continue
        
        if in_action_table and any("パッケージ" in v or "クラス" in v for v in vals):
            # サブヘッダー行（スキップ or マージ）
            continue
        
        if in_action_table:
            # データ行（先頭がNoで始まる or 数値）
            non_empty = [v for v in vals if v]
            if non_empty:
                action_rows.append(vals)
    
    # 基本情報出力
    if basic_info:
        lines.append("### 基本情報\n")
        for k, v in basic_info.items():
            lines.append(f"- **{k}**: {v}")
        lines.append("")
    
    # アクション一覧テーブル
    if action_headers and action_rows:
        lines.append("### 処理一覧\n")
        # ヘッダーを整理（Noneを除去してコンパクトに）
        clean_headers = [v if v else "" for v in action_headers]
        # 有意な列だけを抽出
        used_cols = set()
        for row in action_rows:
            for i, v in enumerate(row):
                if v:
                    used_cols.add(i)
        used_cols = sorted(used_cols)
        
        if used_cols:
            h = [clean_headers[i] if i < len(clean_headers) else "" for i in used_cols]
            r = [[row[i] if i < len(row) else "" for i in used_cols] for row in action_rows]
            lines.append(make_md_table(h, r))
    
    return "\n".join(lines)


def parse_design_overview(wb: openpyxl.Workbook, sheet_name: str) -> str:
    """設計概要シートをMarkdownに変換する。"""
    ws = wb[sheet_name]
    # シート名からタイトルを生成
    title = sheet_name.replace("_", " ").strip()
    lines = [f"## {title}\n"]
    
    # テキストセルを収集
    for row in ws.iter_rows(min_row=1, values_only=False):
        vals = row_values(row)
        if is_empty_row(vals):
            continue
        # 非空セルを結合して出力
        non_empty = [v for v in vals if v]
        if non_empty:
            text = "　".join(non_empty)
            # セクションヘッダー（■で始まる）
            if text.startswith("■"):
                lines.append(f"\n### {text}\n")
            else:
                lines.append(text)
    
    return "\n".join(lines)


def parse_table_sheet(wb: openpyxl.Workbook, sheet_name: str, title: str) -> str:
    """汎用テーブルシートをMarkdownに変換する。
    ヘッダー行を自動検出し、データ行をテーブル化する。"""
    ws = wb[sheet_name]
    lines = [f"## {title}\n"]
    
    all_rows = []
    for row in ws.iter_rows(min_row=1, values_only=False):
        vals = row_values(row)
        if not is_empty_row(vals):
            all_rows.append(vals)
    
    if not all_rows:
        lines.append("> ※ データが見つかりませんでした。")
        return "\n".join(lines)
    
    # ヘッダー候補を探す: 最初の「No」を含む行、または最初の非空行
    header_idx = 0
    sub_header_idx = -1
    for i, row in enumerate(all_rows):
        if any("No" in v for v in row if v):
            header_idx = i
            # 直後の行がサブヘッダーか確認
            if i + 1 < len(all_rows):
                next_row = all_rows[i + 1]
                # サブヘッダー判定：数値Noで始まらない & テキストが多い
                first_val = next((v for v in next_row if v), "")
                try:
                    int(first_val)
                except (ValueError, TypeError):
                    if first_val and not first_val.isdigit():
                        sub_header_idx = i + 1
            break
    
    # ヘッダーを構築（メインヘッダーとサブヘッダーをマージ）
    headers = all_rows[header_idx]
    if sub_header_idx >= 0:
        sub = all_rows[sub_header_idx]
        merged = []
        for j in range(max(len(headers), len(sub))):
            h = headers[j] if j < len(headers) else ""
            s = sub[j] if j < len(sub) else ""
            if h and s:
                merged.append(f"{h}({s})")
            elif h:
                merged.append(h)
            else:
                merged.append(s)
        headers = merged
        data_start = sub_header_idx + 1
    else:
        data_start = header_idx + 1
    
    # タイトル行の前にテキストがあれば出力
    for i in range(header_idx):
        non_empty = [v for v in all_rows[i] if v]
        if non_empty:
            text = "　".join(non_empty)
            if "画面項目" in text or "定義" in text or sheet_name in text:
                continue  # シートタイトル行をスキップ
            lines.append(text + "\n")
    
    # 有意な列のみ抽出
    data_rows = all_rows[data_start:]
    used_cols = set()
    for row in [headers] + data_rows:
        for i, v in enumerate(row):
            if v:
                used_cols.add(i)
    used_cols = sorted(used_cols)
    
    if used_cols and data_rows:
        h = [headers[i] if i < len(headers) else f"列{i+1}" for i in used_cols]
        r = [[row[i] if i < len(row) else "" for i in used_cols] for row in data_rows]
        lines.append(make_md_table(h, r))
    else:
        lines.append("> ※ テーブルデータを抽出できませんでした。")
    
    return "\n".join(lines)


def parse_screen_transition(wb: openpyxl.Workbook, sheet_name: str, xlsx_path: str) -> str:
    """画面遷移図シートをMarkdownに変換する。
    extract_excel_shapes.py の機能を使用して図形からMermaidを生成する。"""
    lines = [f"## 画面遷移図（{sheet_name}）\n"]
    
    # 図形抽出スクリプトをインポート
    script_dir = os.path.dirname(os.path.abspath(__file__))
    shapes_script = os.path.join(script_dir, "extract_excel_shapes.py")
    
    if os.path.exists(shapes_script) and etree is not None:
        # extract_excel_shapes のモジュールをインポート
        import importlib.util
        spec = importlib.util.spec_from_file_location("extract_excel_shapes", shapes_script)
        shapes_mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(shapes_mod)
        
        try:
            result = shapes_mod.analyze_excel_file(xlsx_path, sheet_name)
            
            if sheet_name in result.get("drawings", {}):
                data = result["drawings"][sheet_name]
                
                # 画面一覧
                action_kw = ['ボタン', '押下', 'リンク', 'アイコン', 'クリック', '出力']
                screen_shapes = [
                    s for s in data['shapes']
                    if s['type'] == 'shape' and s['text'].strip()
                    and not any(kw in s['text'] for kw in action_kw)
                    and 'callout' not in s.get('preset', '').lower()
                ]
                if screen_shapes:
                    lines.append("### 画面一覧（図形から抽出）\n")
                    s_headers = ["No", "画面名", "図形タイプ"]
                    s_rows = [[str(i), s["text"], s["preset"] or "-"] for i, s in enumerate(screen_shapes, 1)]
                    lines.append(make_md_table(s_headers, s_rows))
                    lines.append("")
                
                # Mermaid遷移図
                lines.append("### 遷移図\n")
                lines.append(data["mermaid"])
                lines.append("")
                
                # 遷移定義テーブル
                lines.append("### 遷移定義テーブル\n")
                lines.append(data["transition_table"])
            else:
                lines.append("> ※ このシートにはDrawing（図形）データが見つかりませんでした。")
        except Exception as e:
            lines.append(f"> ※ 図形抽出エラー: {e}")
    else:
        if etree is None:
            lines.append("> ※ lxmlがインストールされていないため、図形の抽出ができません。")
            lines.append("> `pip install lxml` を実行してください。")
        else:
            lines.append("> ※ extract_excel_shapes.py が見つかりません。")
    
    return "\n".join(lines)


def parse_excel_layout(wb: openpyxl.Workbook, sheet_name: str) -> str:
    """Excelレイアウトシートをテキスト形式で出力する。"""
    ws = wb[sheet_name]
    title = sheet_name.strip()
    lines = [f"## {title}\n"]
    
    lines.append("> ※ Excelレイアウトシートです。セルの配置・書式情報を含むため、テーブル形式で概要を出力します。\n")
    
    all_rows = []
    for row in ws.iter_rows(min_row=1, max_row=50, values_only=False):
        vals = row_values(row)
        if not is_empty_row(vals):
            all_rows.append(vals)
    
    if not all_rows:
        lines.append("> ※ データが見つかりませんでした。")
        return "\n".join(lines)
    
    # 有意な列のみ抽出
    used_cols = set()
    for row in all_rows:
        for i, v in enumerate(row):
            if v:
                used_cols.add(i)
    used_cols = sorted(used_cols)
    
    if used_cols:
        # 最初の行をヘッダーとして使用
        h = [all_rows[0][i] if i < len(all_rows[0]) and all_rows[0][i] else f"列{i+1}" for i in used_cols]
        r = [[row[i] if i < len(row) else "" for i in used_cols] for row in all_rows[1:]]
        lines.append(make_md_table(h, r))
    
    return "\n".join(lines)


# =====================================================================
#  メイン変換処理
# =====================================================================

def convert_excel_to_md(xlsx_path: str, output_path: Optional[str] = None) -> str:
    """
    Excelファイルを解析してMarkdown基本設計書に変換する。
    
    Args:
        xlsx_path: 入力Excelファイルパス
        output_path: 出力Markdownファイルパス（Noneの場合はコンソール出力）
    
    Returns:
        生成されたMarkdown文字列
    """
    print(f"解析開始: {xlsx_path}")
    
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet_names = wb.sheetnames
    classified = classify_sheets(sheet_names)
    
    print(f"シート数: {len(sheet_names)}")
    print(f"分類結果:")
    for cat, names in classified.items():
        print(f"  {cat}: {names}")
    
    # 基本情報取得
    basic_info = extract_basic_info(wb, sheet_names)
    
    # Markdown構築
    md_parts = []
    
    # タイトル
    func_name = basic_info.get("業務機能名", os.path.basename(xlsx_path).replace(".xlsx", ""))
    doc_id = basic_info.get("管理No", "")
    title = f"# 業務機能設計書: {func_name}"
    if doc_id:
        title += f"\n\n> 管理No: {doc_id}"
    if basic_info.get("機能概要"):
        title += f" | 機能概要: {basic_info['機能概要']}"
    md_parts.append(title)
    md_parts.append("")
    
    # 目次
    md_parts.append("## 目次\n")
    toc_index = 1
    for cat, names in classified.items():
        for name in names:
            md_parts.append(f"{toc_index}. [{name}](#{name.replace(' ', '-')})")
            toc_index += 1
    md_parts.append("")
    md_parts.append("---\n")
    
    # 各シートの変換
    for cat, names in classified.items():
        for sheet_name in names:
            print(f"  処理中: {sheet_name} ({cat})")
            try:
                if cat == "change_history":
                    md_parts.append(parse_change_history(wb, sheet_name))
                elif cat == "spec_overview":
                    md_parts.append(parse_spec_overview(wb, sheet_name))
                elif cat == "design_overview":
                    md_parts.append(parse_design_overview(wb, sheet_name))
                elif cat == "screen_items":
                    md_parts.append(parse_table_sheet(wb, sheet_name, f"画面項目定義"))
                elif cat == "screen_state":
                    md_parts.append(parse_table_sheet(wb, sheet_name, f"画面状態定義"))
                elif cat == "screen_check":
                    md_parts.append(parse_table_sheet(wb, sheet_name, f"画面項目チェック定義"))
                elif cat == "screen_source":
                    md_parts.append(parse_table_sheet(wb, sheet_name, f"画面取得元定義"))
                elif cat == "screen_transition":
                    md_parts.append(parse_screen_transition(wb, sheet_name, xlsx_path))
                elif cat == "excel_transfer":
                    md_parts.append(parse_table_sheet(wb, sheet_name, f"Excel移送表: {sheet_name}"))
                elif cat == "excel_layout":
                    md_parts.append(parse_excel_layout(wb, sheet_name))
                elif cat == "process_detail":
                    md_parts.append(parse_table_sheet(wb, sheet_name, f"処理詳細設計: {sheet_name}"))
                elif cat == "table_mapping":
                    md_parts.append(parse_table_sheet(wb, sheet_name, f"テーブル項目マッピング: {sheet_name}"))
                else:
                    # その他シート：汎用テーブル or テキスト
                    md_parts.append(parse_design_overview(wb, sheet_name))
            except Exception as e:
                md_parts.append(f"## {sheet_name}\n\n> ※ 解析エラー: {e}\n")
            
            md_parts.append("\n---\n")
    
    wb.close()
    
    # 結合
    md_content = "\n".join(md_parts)
    
    # 出力
    if output_path:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(md_content)
        print(f"\n出力完了: {output_path}")
    
    return md_content


def main():
    parser = argparse.ArgumentParser(
        description="Excel業務機能設計書をMarkdown基本設計書に変換する",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--input", required=True, help="入力Excelファイルパス")
    parser.add_argument("--output", default=None, help="出力Markdownファイルパス（省略時はコンソール出力）")
    
    args = parser.parse_args()
    
    if not os.path.exists(args.input):
        print(f"エラー: ファイルが見つかりません: {args.input}", file=sys.stderr)
        sys.exit(1)
    
    md = convert_excel_to_md(args.input, args.output)
    
    if not args.output:
        print("\n" + "=" * 70)
        print(md)


if __name__ == "__main__":
    main()