#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
screenshot_to_excel.py
读取指定文件夹下的测试截图图片，将图片整理到一个 Excel（xlsx）中。

使用方式：
    python screenshot_to_excel.py --input-dir "D:\\test_screenshots" --output-xlsx "D:\\test_screenshots\\screenshots.xlsx"

依赖安装：
    pip install openpyxl pillow
"""

import argparse
import io
import os
import shutil
import sys
import tempfile
from pathlib import Path
from typing import List, Tuple

# 修复控制台编码问题（Windows 环境下 cp932/gbk 无法输出中文简体字符）
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
if sys.stderr.encoding and sys.stderr.encoding.lower() not in ("utf-8", "utf8"):
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

# 支持的图片后缀名
SUPPORTED_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp"}

# 图片缩放最大宽度（像素）
MAX_IMAGE_WIDTH = 600

# Excel 列宽设置
COL_WIDTH_A = 8    # 序号列
COL_WIDTH_B = 40   # 文件名列
COL_WIDTH_C = 70   # 图片列

# Excel 默认行高（像素 -> 磅的近似值，1磅 ≈ 1.333px）
DEFAULT_ROW_HEIGHT = 160  # 磅


def collect_images(input_dir: str) -> List[Path]:
    """
    扫描指定目录下符合后缀名的图片文件，按文件名排序后返回路径列表。

    Args:
        input_dir: 截图图片所在文件夹路径

    Returns:
        排序后的图片路径列表
    """
    input_path = Path(input_dir)
    if not input_path.exists():
        print(f"错误：输入目录不存在：{input_dir}", file=sys.stderr)
        sys.exit(1)

    if not input_path.is_dir():
        print(f"错误：输入路径不是目录：{input_dir}", file=sys.stderr)
        sys.exit(1)

    images = []
    for f in input_path.iterdir():
        if f.is_file() and f.suffix.lower() in SUPPORTED_EXTENSIONS:
            images.append(f)

    # 按文件名字符串排序，保证稳定输出
    images.sort(key=lambda p: p.name)

    if not images:
        print(f"错误：目录下没有找到匹配的图片文件：{input_dir}", file=sys.stderr)
        print(f"支持的格式：{', '.join(SUPPORTED_EXTENSIONS)}", file=sys.stderr)
        sys.exit(1)

    return images


def resize_image(src_path: Path, dst_path: Path, max_width: int = MAX_IMAGE_WIDTH) -> Tuple[int, int]:
    """
    将图片按固定最大宽度等比缩放，保存到目标路径。

    Args:
        src_path: 源图片路径
        dst_path: 缩放后图片保存路径
        max_width: 最大宽度（像素）

    Returns:
        缩放后的 (宽度, 高度) 元组

    Raises:
        Exception: 图片读取或处理失败时抛出
    """
    img = PILImage.open(src_path)

    # 如果图片有透明通道且需要保存为 JPEG，先转换
    original_width, original_height = img.size

    if original_width > max_width:
        ratio = max_width / original_width
        new_width = max_width
        new_height = int(original_height * ratio)
        img = img.resize((new_width, new_height), PILImage.LANCZOS)
    else:
        new_width, new_height = original_width, original_height

    # 统一保存为 PNG 格式，避免透明通道和格式兼容问题
    # 如果原图是 RGBA 模式，保持 PNG；否则也用 PNG 保存
    if img.mode == "RGBA":
        img.save(dst_path, format="PNG")
    else:
        # 转为 RGB 再保存为 PNG（处理一些特殊模式如 P, LA 等）
        if img.mode not in ("RGB", "L"):
            img = img.convert("RGB")
        img.save(dst_path, format="PNG")

    return new_width, new_height


def build_workbook(image_paths: List[Path], output_xlsx: str) -> None:
    """
    创建 Excel 工作簿，将图片逐行插入到 Screenshots 工作表中。

    Args:
        image_paths: 图片路径列表（已排序）
        output_xlsx: 输出 Excel 文件路径
    """
    # 确保输出目录存在
    output_path = Path(output_xlsx)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # 创建临时目录用于存放缩放后的图片
    tmp_dir = output_path.parent / ".tmp_images"
    tmp_dir.mkdir(exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Screenshots"

    # 设置列宽
    ws.column_dimensions["A"].width = COL_WIDTH_A
    ws.column_dimensions["B"].width = COL_WIDTH_B
    ws.column_dimensions["C"].width = COL_WIDTH_C

    # 写入表头
    header_row = 1
    ws.cell(row=header_row, column=1, value="序号")
    ws.cell(row=header_row, column=2, value="文件名")
    ws.cell(row=header_row, column=3, value="图片")

    # 表头行设置加粗
    from openpyxl.styles import Font, Alignment
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 4):
        cell = ws.cell(row=header_row, column=col)
        cell.font = header_font
        cell.alignment = header_alignment

    ws.row_dimensions[header_row].height = 25

    # 数据行从第2行开始
    data_start_row = 2
    success_count = 0
    skip_count = 0

    for idx, img_path in enumerate(image_paths, start=1):
        row = data_start_row + idx - 1

        try:
            # 缩放图片到临时目录
            tmp_img_path = tmp_dir / f"{idx:04d}_{img_path.stem}.png"
            new_w, new_h = resize_image(img_path, tmp_img_path)

            # 写入序号和文件名
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=img_path.name)

            # 设置单元格对齐方式
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="center")

            # 插入图片到 C 列
            xl_img = XlImage(str(tmp_img_path))
            # 设置图片在 Excel 中的显示大小（openpyxl 使用 EMU 或像素）
            xl_img.width = new_w
            xl_img.height = new_h
            ws.add_image(xl_img, f"C{row}")

            # 根据图片高度动态设置行高
            # 像素转磅：1磅 ≈ 1.333像素，所以行高(磅) ≈ 像素 / 1.333
            row_height_points = new_h / 1.333 + 10  # 额外留 10 磅的边距
            if row_height_points < 30:
                row_height_points = 30
            ws.row_dimensions[row].height = row_height_points

            success_count += 1
            print(f"  [{idx}/{len(image_paths)}] 已处理：{img_path.name} ({new_w}x{new_h})")

        except Exception as e:
            # 跳过损坏或无法读取的图片，继续处理
            print(f"  [{idx}/{len(image_paths)}] 跳过（错误）：{img_path.name} - {e}", file=sys.stderr)
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=f"{img_path.name} (处理失败)")
            ws.row_dimensions[row].height = 30
            skip_count += 1
            continue

    # 保存 Excel
    wb.save(output_xlsx)

    # 清理临时目录
    try:
        shutil.rmtree(tmp_dir)
    except Exception as e:
        print(f"警告：清理临时目录失败：{e}", file=sys.stderr)

    print(f"\n完成！")
    print(f"  成功处理：{success_count} 张图片")
    if skip_count > 0:
        print(f"  跳过：{skip_count} 张图片")
    print(f"  输出文件：{output_xlsx}")


def main():
    parser = argparse.ArgumentParser(
        description="将截图文件夹中的图片整理到 Excel 中",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例：
  python screenshot_to_excel.py --input-dir "D:\\test_screenshots"
  python screenshot_to_excel.py --input-dir "./images" --output-xlsx "./output/screenshots.xlsx"

支持的图片格式：png, jpg, jpeg, webp
        """
    )

    parser.add_argument(
        "--input-dir",
        required=True,
        help="截图图片所在文件夹路径（必填）"
    )

    parser.add_argument(
        "--output-xlsx",
        default=None,
        help="输出 Excel 路径（可选，默认在输入目录下生成 screenshots.xlsx）"
    )

    args = parser.parse_args()

    input_dir = args.input_dir
    output_xlsx = args.output_xlsx

    # 如果未指定输出路径，默认在输入目录下生成
    if output_xlsx is None:
        output_xlsx = str(Path(input_dir) / "screenshots.xlsx")

    print(f"输入目录：{input_dir}")
    print(f"输出文件：{output_xlsx}")
    print(f"支持格式：{', '.join(SUPPORTED_EXTENSIONS)}")
    print(f"最大图片宽度：{MAX_IMAGE_WIDTH}px")
    print()

    # 收集图片
    image_paths = collect_images(input_dir)
    print(f"找到 {len(image_paths)} 张图片，开始处理...\n")

    # 构建工作簿
    build_workbook(image_paths, output_xlsx)


if __name__ == "__main__":
    main()